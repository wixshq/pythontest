"""
=============
Author: whq
Time: 2022-08-30 18:35


"""
import os
from openpyxl import load_workbook




class HandleExcel:

    def __init__(self,file_fir,sheet_name):
        # 实例化wb对象
        self.wb = load_workbook(file_fir)
        # 实例化sh对象
        self.sh = self.wb[sheet_name]


    def __read_titles(self):
        titles = []
        # 获取表头的值，组装成一个list
        for item in list(self.sh.rows)[0]:
            titles.append(item.value)
        return titles

    def read_all_datas(self):
        # 从Excel第二行开始循环，将单元格的值组成list，然后通过zip函数和表头行组成dict，最后追加到最终的数据list
        data_list = []
        titles = self.__read_titles()
        print(titles)
        for item in list(self.sh.rows)[1:]:
            item_list = []
            for i in item:
                item_list.append(i.value)
            res = dict(zip(titles,item_list))
            # eval函数去掉''
            # res["check"] = eval(res["check"])
            data_list.append(res)
        return data_list

    def excel_close(self):
        self.wb.close()

if __name__ == '__main__':
    file_fir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../login_cases.xlsx")
    print(file_fir)
    ex = HandleExcel(file_fir, 'login')
    datas = ex.read_all_datas()
    ex.excel_close()
    print(datas)

# 获取单元格的值
# cel = sh.cell(2,3)
# print(cel.value)
# print(sh)
# print(sh.max_row)
# print(sh.max_column)


# 将Excel的数据组装到list






#
#     value_dict = {}
#     for index in range(len(item)):
#         value_dict[titles[index]] = item[index].value
#     data_list.append(value_dict)
# print(data_list)





# for c in range(2,sh.max_row+1):
#     value_dict = {}
#     for r in range(1,sh.max_column+1):
#         print(sh.cell(c,r).value)
#         value_dict[titles[r]] = sh.cell(c,r).value
#     print(value_dict)


