"""
=============
Author: whq
Time: 2022-08-30 18:35


"""
from openpyxl import load_workbook
import  os

# 获取目录下的测试数据路径
file_fir = os.path.join(os.path.dirname(os.path.abspath(__file__)),"login_cases.xlsx")
print(file_fir)

#实例化wb对象
wb = load_workbook(file_fir)
# 实例化sh对象
sh = wb['login']

# 获取单元格的值
# cel = sh.cell(2,3)
# print(cel.value)
# print(sh)
# print(sh.max_row)
# print(sh.max_column)


# 将Excel的数据组装到list
data_list = []
titles = []
# 获取表头的值，组装成一个list
for item in list(sh.rows)[0]:
    titles.append(item.value)

# 从Excel第二行开始循环，将单元格的值组成list，然后通过zip函数和表头行组成dict，最后追加到最终的数据list
for item in list(sh.rows)[1:]:
    item_list = []
    for i in item:
        item_list.append(i.value)
    res = dict(zip(titles,item_list))
    # eval函数去掉''
    res["check"] = eval(res["check"])
    data_list.append(res)
print(data_list)


#
#     value_dict = {}
#     for index in range(len(item)):
#         value_dict[titles[index]] = item[index].value
#     data_list.append(value_dict)
# print(data_list)





# for c in range(2,sh.max_row+1):
#     value_dict = {}
#     for r in range(1,sh.max_column+1):
#         print(sh.cell(c,r).value)
#         value_dict[titles[r]] = sh.cell(c,r).value
#     print(value_dict)


